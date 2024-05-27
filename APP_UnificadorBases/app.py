from flask import Flask, render_template, request, session, redirect, url_for, send_file
import pandas as pd
import os
import openpyxl
import xlrd

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta'

DIRETORIO_UPLOADS = 'uploads'
EXTENSOES_PERMITIDAS = {'csv', 'xlsx', 'xls'}

app.config['DIRETORIO_UPLOADS'] = DIRETORIO_UPLOADS
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limite de 16 MB para uploads

# Certifique-se de que o direório de uploads exista
os.makedirs(DIRETORIO_UPLOADS, exist_ok=True)

# Definindo a variável global
arquivo_gerado = None

def extensoes_permitidas(arquivos):
    """Valida se as extensões dos arquivos são permitidas."""
    for arquivo in arquivos:
        if '.' not in arquivo.filename:
            return False
        extensao = arquivo.filename.rsplit('.', 1)[1].lower()
        if extensao not in EXTENSOES_PERMITIDAS:
            return False
    return True

# Função que verifica somente se o arquivo CSV existe
def verificar_padrao_arquivo_csv(arquivo):
    """Verifica se o arquivo é um CSV (sempre retorna True para CSV)."""
    return True

# Função que verificar se o arquivo XLS tem mais de uma planilha e se nas 10 primeiras linhas existe mais de uma célula preenchida.
def verificar_padrao_arquivo_xls(arquivo):
    # Abrindo o arquivo Excel
    workbook = xlrd.open_workbook(arquivo)
    
    # Verifica o número de planilhas
    if len(workbook.sheet_names()) != 1:
        return False
    
    # Obtém a primeira planilha
    sheet = workbook.sheet_by_index(0)
    
    # Obtém o número de linhas e colunas
    num_rows = sheet.nrows
    num_cols = sheet.ncols
    
    # Define o número máximo de células preenchidas nas primeiras 10 linhas
    max_cells_filled = 1
    
    # Verifica as primeiras 10 linhas
    for row in range(min(num_rows, 10)):
        cells_filled = sum(1 for col in range(num_cols) if sheet.cell_value(row, col))
        if cells_filled > max_cells_filled:
            # Se encontrar mais de uma célula preenchida, aceita o arquivo
            return True
    
    # Se não encontrar mais de uma célula preenchida nas primeiras 10 linhas, recusa o arquivo
    return False

# Função que verificar se o arquivo XLSX tem mais de uma planilha e se nas 10 primeiras linhas existe mais de uma célula preenchida.

def verificar_padrao_arquivo_xlsx(arquivo):
    # Abrindo o arquivo Excel
    workbook = openpyxl.load_workbook(arquivo)
    
    # Verifica o número de planilhas
    if len(workbook.sheetnames) != 1:
        return False
    
    # Obtém a primeira planilha
    sheet = workbook.active
    
    # Define o número máximo de células preenchidas nas primeiras 10 linhas
    max_cells_filled = 1
    
    # Verifica as primeiras 10 linhas
    for row in range(1, min(sheet.max_row, 11)):
        cells_filled = sum(1 for col in range(1, sheet.max_column + 1) if sheet.cell(row=row, column=col).value)
        if cells_filled > max_cells_filled:
            # Se encontrar mais de uma célula preenchida, aceita o arquivo
            return True
    
    # Se não encontrar mais de uma célula preenchida nas primeiras 10 linhas, recusa o arquivo
    return False

@app.route('/', methods=['POST'])
def upload_arquivo():
    """Rota para lidar com o envio de arquivos."""
    if 'arquivos' not in request.files:
        mensagem_erro = 'Nenhum arquivo enviado. Por favor, selecione um arquivo e tente novamente.'
        return redirect(url_for('formulario_upload', erro=mensagem_erro))

    arquivos = request.files.getlist('arquivos')  # Obter a lista de arquivos enviados

    print("Headers da requisição:", request.headers)
    print("Content-Type:", request.content_type)

    if not extensoes_permitidas(arquivos):
        mensagem_erro = 'Extensão de arquivo não permitida. Por favor, envie arquivos CSV, XLSX ou XLS.'
        return redirect(url_for('formulario_upload', erro=mensagem_erro))

    for arquivo in arquivos:
        if arquivo.filename == '':
            mensagem_erro = 'Nenhum arquivo selecionado. Por favor, selecione um arquivo e tente novamente.'
            return redirect(url_for('formulario_upload', erro=mensagem_erro))

        nome_arquivo = arquivo.filename
        caminho_arquivo = os.path.join(app.config['DIRETORIO_UPLOADS'], nome_arquivo)
        print(f"Salvando arquivo: {caminho_arquivo}")
        arquivo.save(caminho_arquivo)

        arquivos_enviados = session.get('arquivos_enviados', [])
        arquivos_enviados.append(nome_arquivo)
        session['arquivos_enviados'] = arquivos_enviados

        # Verifica se o arquivo é um CSV
        if nome_arquivo.lower().endswith('.csv'):
            if not verificar_padrao_arquivo_csv(caminho_arquivo):
                mensagem_erro = f'Conteúdo incorreto!<br>Caracteres permitidos [,] ou [;].'
                limpar()
                return redirect(url_for('formulario_upload', erro=mensagem_erro))
        # Verifica se o arquivo é uma tabela XLS
        elif nome_arquivo.lower().endswith('.xls'):
            if not verificar_padrao_arquivo_xls(caminho_arquivo):
                mensagem_erro = f'Conteúdo incorreto!<br>Somente tabela.'
                limpar()
                return redirect(url_for('formulario_upload', erro=mensagem_erro))
        # Verifica se o arquivo é uma tabela XLSX
        elif nome_arquivo.lower().endswith('.xlsx'):
            if not verificar_padrao_arquivo_xlsx(caminho_arquivo):
                mensagem_erro = f'Conteúdo incorreto!<br>Somente tabela.'
                limpar()
                return redirect(url_for('formulario_upload', erro=mensagem_erro))

    session['sucesso'] = "Enviado com sucesso!"
    return render_template('upload.html', arquivos_enviados=arquivos_enviados, sucesso=session['sucesso'])

@app.route('/processar')
def processar_arquivos():
    global arquivo_gerado
    """Rota para processar os arquivos enviados."""
    try:
        arquivos = os.listdir(app.config['DIRETORIO_UPLOADS'])
        if not arquivos:
            mensagem_erro = 'Nenhum arquivo para processar.'
            return redirect(url_for('formulario_upload', erro=mensagem_erro))

        tabela_consolidada = pd.DataFrame()

        for arquivo in arquivos:
            caminho_arquivo = os.path.join(app.config['DIRETORIO_UPLOADS'], arquivo)
            if arquivo.endswith('.csv'):
                try:
                    tabela_vendas = pd.read_csv(caminho_arquivo, encoding='utf-8')
                except UnicodeDecodeError:
                    tabela_vendas = pd.read_csv(caminho_arquivo, encoding='latin1')
                except pd.errors.ParserError:
                    try:
                        tabela_vendas = pd.read_csv(caminho_arquivo, encoding='utf-8', delimiter=';')
                    except UnicodeDecodeError:
                        tabela_vendas = pd.read_csv(caminho_arquivo, encoding='latin1', delimiter=';')
            elif arquivo.endswith('.xlsx'):
                tabela_vendas = pd.read_excel(caminho_arquivo)
            elif arquivo.endswith('.xls'):
                tabela_vendas = pd.read_excel(caminho_arquivo, engine='xlrd')
            else:
                continue

            print(f"Arquivo lido com sucesso: {arquivo}")
            tabela_consolidada = pd.concat([tabela_consolidada, tabela_vendas])

        nome_arquivo_saida = os.path.join(app.config['DIRETORIO_UPLOADS'], 'arquivos.xlsx')
        arquivo_gerado = nome_arquivo_saida
        tabela_consolidada.to_excel(nome_arquivo_saida, index=False)

        return redirect(url_for('formulario_upload', download_file=nome_arquivo_saida))
    except Exception as e:
        mensagem_erro = f'Erro ao processar arquivos: {str(e)}'
        print(mensagem_erro)  # Log de erro
        return redirect(url_for('formulario_upload', erro=mensagem_erro))

def limpar():
    """Limpa a pasta de uploads e a lista de arquivos enviados."""
    for arquivo in os.listdir(app.config['DIRETORIO_UPLOADS']):
        os.remove(os.path.join(app.config['DIRETORIO_UPLOADS'], arquivo))
    session['arquivos_enviados'] = []  # Define a lista de arquivos enviados como vazia
    global arquivo_gerado
    arquivo_gerado = None

@app.route('/', methods = ['GET'])
def formulario_upload():
    global arquivo_gerado
    arquivos_enviados = session.get('arquivos_enviados')
    # Verifica se há um arquivo para download
    download_file = request.args.get('download_file')

    if download_file:
        return render_template('upload.html', sucesso="Arquivo Gerado!", arquivo_gerado=arquivo_gerado, arquivos_enviados=arquivos_enviados)
    else:
        # Renderiza o formulário de upload com possível mensagem de erro
        erro = request.args.get('erro')
        return render_template('upload.html', erro=erro)

@app.route('/download')
def download_base():
    global arquivo_gerado
    if arquivo_gerado is not None:
        enviar = send_file(arquivo_gerado, as_attachment=True)
        limpar()
        return enviar
    else:
        # Renderiza o formulário de upload com possível mensagem de erro
        erro = "Não existe arquivo para Download!"
        return render_template('upload.html', erro=erro)
    
@app.route('/restaurar')
def restaurar():
    global arquivo_gerado
    if arquivo_gerado is not None:
        limpar()
        return redirect(url_for('upload_arquivo'))
    else:
        # Renderiza o formulário de upload com possível mensagem de erro
        limpar()
        return redirect(url_for('upload_arquivo'))

if __name__ == '__main__':
    app.run(debug=True)
